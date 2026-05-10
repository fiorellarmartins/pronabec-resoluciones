#!/usr/bin/env python3
"""
Export person-level LLM classifications to a formatted Excel file.
"""
import csv, json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_IN  = "/Users/fiorellaramirez/Desktop/pronabec_datos/llm_classification/persons_final.csv"
JSONL   = "/Users/fiorellaramirez/Desktop/pronabec_datos/llm_classification/results.jsonl"
OUT     = "/Users/fiorellaramirez/Desktop/pronabec_datos/clasificacion_pronabec.xlsx"

# ── load data ─────────────────────────────────────────────────────────────────
persons = list(csv.DictReader(open(CSV_IN, encoding="utf-8")))

CLASS_LABEL = {
    "cumplió":        "Cumplió compromiso",
    "devolución":     "Devolución",
    "incumplimiento": "Incumplimiento",
    "aplazamiento":   "Aplazamiento",
    "improcedente":   "Improcedente",
    "nulidad":        "Nulidad",
    "otro":           "Otro",
}
SUB_LABEL = {
    "acreditación":           "Acreditación procedente",
    "acreditación_parcial":   "Acreditación parcial",
    "cumplimiento_directo":   "Cumplimiento directo",
    "abandono":               "Abandono",
    "bajo_rendimiento":       "Bajo rendimiento",
    "renuncia":               "Renuncia voluntaria",
    "no_aceptación":          "No aceptó beca",
    "pérdida_otros":          "Pérdida / otros",
    "declaración":            "Declaración de incumplimiento",
    "recurso_improcedente":   "Recurso improcedente",
    "motivos_académicos":     "Motivos académicos",
    "motivos_laborales":      "Motivos laborales",
    "motivos_salud":          "Motivos de salud",
    "otros_motivos":          "Otros motivos",
    "aplazamiento":           "Aplazamiento denegado",
    "ampliación":             "Ampliación denegada",
    "exoneración":            "Exoneración denegada",
    "acreditación_denegada":  "Acreditación denegada",
    "otros":                  "Otros",
    "fundado_recurso":        "Recurso fundado",
    "sin_efecto":             "Resolución sin efecto",
    "admisión":               "Admisión",
    "viaje":                  "Viaje / retorno",
    "modificación":           "Modificación",
    "excepción":              "Excepción",
    "admin":                  "Admin / trámites",
    "improcedente":           "Improcedente",
}
BECA_LABEL = {
    "beca_presidente":   "Beca Presidente de la República",
    "beca_bicentenario": "Beca Generación Bicentenario",
}
CLASS_COLORS = {
    "cumplió":        "C6EFCE",
    "devolución":     "FFC7CE",
    "incumplimiento": "FF9999",
    "aplazamiento":   "E2CFEB",
    "improcedente":   "FFDDC1",
    "nulidad":        "D9D9D9",
    "otro":           "FFFFCC",
}

rows = []
for p in persons:
    cls = p["final_class"]
    sub = p["final_subclass"]
    rows.append({
        "DNI":              p["dni"],
        "Nombre":           p["person"],
        "Beca":             BECA_LABEL.get(p["beca"], p["beca"]),
        "Convocatoria":     p["convocatoria"],
        "Clasificación":    CLASS_LABEL.get(cls, cls),
        "Subclasificación": SUB_LABEL.get(sub, sub),
        "Regex coincide":   "Sí" if str(p["llm_validated"]).lower() == "true" else "No",
        "Resumen":          p["summary"],
        "Archivo":          p["filename"],
    })

df = pd.DataFrame(rows)
df = df.sort_values(["Clasificación", "Convocatoria", "Nombre"]).reset_index(drop=True)

# ── write with openpyxl ───────────────────────────────────────────────────────
with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Personas", index=False)

    # summary sheet
    from collections import Counter
    summary_rows = []
    for cls in ["cumplió","aplazamiento","devolución","incumplimiento","improcedente","nulidad","otro"]:
        subset = [p for p in persons if p["final_class"] == cls]
        if not subset: continue
        summary_rows.append({
            "Clasificación": CLASS_LABEL.get(cls, cls),
            "N personas": len(subset),
            "% del total": f"{len(subset)/len(persons)*100:.1f}%",
        })
    pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Resumen", index=False)

# ── formatting ────────────────────────────────────────────────────────────────
wb = load_workbook(OUT)

for sheet_name in ["Personas", "Resumen"]:
    ws = wb[sheet_name]

    # header row
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # data rows
    cls_col = None
    for i, col in enumerate(df.columns, 1):
        if col == "Clasificación":
            cls_col = i
            break

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cls_val = ws.cell(row=row_idx, column=cls_col).value if cls_col else ""
        color = None
        for k, v in CLASS_LABEL.items():
            if v == cls_val:
                color = CLASS_COLORS.get(k)
                break
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(cell.column == len(df.columns)))
            if color:
                cell.fill = PatternFill("solid", fgColor=color)

    # column widths
    col_widths = {
        "DNI": 12, "Nombre": 32, "Beca": 30, "Convocatoria": 13,
        "Clasificación": 22, "Subclasificación": 28, "Regex coincide": 14,
        "Resumen": 60, "Archivo": 30,
        # resumen sheet
        "N personas": 12, "% del total": 12,
    }
    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        header = col[0].value
        width  = col_widths.get(header, 18)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  {len(persons)} personas · {len(df)} filas")
